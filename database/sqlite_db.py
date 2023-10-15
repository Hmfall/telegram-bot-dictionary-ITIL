import sqlite3 as sq
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from io import BytesIO
from aiogram.types.input_file import InputFile


def database_name():
	return "ITIL.db"
def dictionary_table_name():
	return "dictionary_object"


def sql_start():
	global base, cur
	base = sq.connect(database_name())
	cur = base.cursor()
	if base:
		print("Data base connected")
	base.execute(f" \
		CREATE TABLE IF NOT EXISTS {dictionary_table_name()}( \
		term TEXT PRIMARY KEY UNIQUE, \
	   term_eng TEXT, \
		definition TEXT)")


async def sql_add_command(state):
	async with state.proxy() as data:
		cur.execute(f"INSERT INTO {dictionary_table_name()} VALUES (?, ?, ?)", tuple(data.values()))
		base.commit()


def is_value_uniqueness(word):
    cur.execute('SELECT * FROM ' + dictionary_table_name() + ' WHERE lower(term) = lower(?)', (word,))
    result = cur.fetchone()
    if result is not None:
        is_uniqueness = result[0]
        return is_uniqueness == 0
    else:
        return True


def excel_initial(excel_initial_file_name):
	workbook = load_workbook(filename=excel_initial_file_name)
	count_added = 0
	count_skipped = 0
	try:
		sheet = workbook.active
		print("Excel is connected...")

		conn = sq.connect(database_name())
		conn.execute("PRAGMA journal_mode=WAL")
		cur = conn.cursor()

		for row in sheet.iter_rows(min_row=0, values_only=True):
			term, term_eng, definition = row
			cur.execute(f"SELECT term FROM {dictionary_table_name()} WHERE term = ?", (term,))
			if cur.fetchone() is None:
				if term is not None and term_eng is not None and definition is not None:
					cur.execute(f"INSERT INTO {dictionary_table_name()} (term, term_eng, definition) VALUES (?, ?, ?)",
					(term.strip(), term_eng.strip(), definition.strip()))
					count_added += 1
					conn.commit()
			else:
				count_skipped += 1
		print("Data:")
		print(f"{count_added} is added")
		print(f"{count_skipped} is skipped")
		return (count_added, count_skipped)
	except KeyError:
		print(f"Sheet not found in file '{excel_initial_file_name}'")
	finally:
		conn.close()


# ITIL.db -> ITIL.xlsx
def export_excel_def():
   base = sq.connect(database_name())
   cur = base.cursor()

   cur.execute(f"SELECT * FROM {dictionary_table_name()}")
   rows = cur.fetchall()

   __wb__ = Workbook()
   ws = __wb__.active
   ws.title = "ITIL.xlsx"

   for row in rows:
      ws.append([str(item) for item in row]) 

	# Excel to BytesIO
   file_buffer = BytesIO()
   __wb__.save(file_buffer)
   file_buffer.seek(0)
   
   filename = "ITIL.xlsx"
   filename_with_timestamp = f"{set_timestamp()}__database__{filename}"
   # __wb__.save(filename_with_timestamp)
   base.close()
   file = InputFile(file_buffer, filename_with_timestamp)
   return file 


def return_total_rows():
	cur.execute(f"SELECT COUNT(*) FROM {dictionary_table_name()}")
	total = cur.fetchone()[0]
	return total


def set_timestamp():
	now = datetime.now()
	timestamp = now.strftime("%d.%m.%Y")
	return timestamp


def clear_table():
	base = sq.connect(database_name())
	cur = base.cursor()
	
	cur.execute(f"SELECT COUNT(*) FROM {dictionary_table_name()}")
	num_rows = cur.fetchone()[0]
	
	if num_rows == 0:
		base.close()
		return False
	
	cur.execute(f"DELETE FROM {dictionary_table_name()}")
	num_deleted = cur.rowcount
	
	base.commit()
	base.close()
	return num_deleted
